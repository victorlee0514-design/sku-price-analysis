import io
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="SKU 价格毛利分析", layout="wide")

# ── 页眉 ──────────────────────────────────────────────────────────────────────
h1, h2, h3 = st.columns([1, 3, 2])
with h1:
    try:
        st.image("carzone_logo.png", width=110)
    except Exception:
        pass
with h2:
    st.markdown("## SKU 价格毛利分析")
    st.caption("底盘件价格体系优化 · 核算价 × 毛利率模拟工具")
with h3:
    st.markdown(
        "<div style='text-align:right;line-height:1.7;font-size:13px;color:#888;padding-top:8px'>"
        "<b>南京新康众 · 供应链底盘组</b><br>制作人：李宇凡<br>"
        "<span style='color:#e74c3c;font-size:12px'>⚠ 仅供底盘组内部使用，注意保护敏感数据安全</span>"
        "</div>",
        unsafe_allow_html=True,
    )
st.divider()

# ── 侧边栏 ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("数据源")
    uploaded = st.file_uploader("上传 Excel 文件（.xlsx）", type=["xlsx"])

# ── 未上传时：使用说明页 ──────────────────────────────────────────────────────
if not uploaded:
    st.markdown("### 欢迎使用 SKU 价格毛利分析看板")
    st.markdown("上传 Excel 文件后，工具将自动识别列结构并开始分析。请先阅读以下使用说明。")

    with st.expander("📂 支持的文件格式", expanded=True):
        st.markdown("""
**推荐格式：SKU 价格导出标准文件**
- 必须包含：`加权平均采购成本`（供应商进货价）和 `核算价`（总部对门店定价）
- 工具会自动识别列名，无需手动调整

**不支持的格式：门店采购明细（如法雷奥/菲罗多采购流水）**
- 此类文件只记录门店向总部的拿货价，不含供应商进货成本，无法计算毛利率
- 上传后工具会自动检测并给出提示
        """)

    with st.expander("🔧 操作步骤", expanded=True):
        st.markdown("""
1. **上传文件**：点击左侧「上传 Excel 文件」，拖拽或选择 SKU 价格导出文件
2. **设置筛选**：在左侧选择品牌和品类（默认显示全部）
3. **调整调价幅度**：拖动「调价幅度」滑块，正数=涨价，负数=降价
4. **设置基准线**：拖动「毛利率基准线」滑块（默认 10%）
5. **调价范围筛选**：勾选「仅调整高于基准线的 SKU」（默认勾选），则当前毛利率 ≤ 基准线的 SKU 不参与调价，核算价维持不变；取消勾选则对所有 SKU 应用调价幅度
6. **读取分析结果**：
   - 顶部摘要卡：一句话总结调价影响（含排除 SKU 数）
   - 三张图：SKU 健康分布 / 毛利率分布对比 / 品类对比
   - 多幅度对比表：同时查看 5 种调价幅度的结果
   - 品类汇总表：各品类均值毛利率和风险 SKU 数
   - SKU 明细表：含最低达标价格和异常标记
7. **导出**：点击「导出带格式 Excel」，下载带颜色的 .xlsx 文件
        """)

    with st.expander("🎨 颜色说明"):
        st.markdown("""
| 颜色 | 含义 |
|------|------|
| 🟢 绿色 | 毛利率 ≥ 基准线，健康 |
| 🟡 黄色 | 毛利率在 8% 到基准线之间，预警 |
| 🔴 红色 | 毛利率 < 8%，危险 |
| ⚠️ 异常标记 | 该 SKU 毛利率偏离所在品类均值超过 1.5 个标准差 |
        """)

    with st.expander("❓ 常见问题"):
        st.markdown("""
**Q：上传后数据显示为空？**
A：检查文件是否含有加权平均采购成本列，且该列有有效数值（非 0 非空）。

**Q：列识别失败怎么办？**
A：工具会弹出手动选列面板，从下拉框中选择对应列即可。

**Q：调价后某些 SKU 毛利率变为负数？**
A：说明该 SKU 当前核算价已低于成本，需要涨价而非降价。

**Q：「最低达标价格」是怎么算的？**
A：最低达标价格 = 采购成本 ÷ (1 - 毛利率基准线)，即刚好达到基准线所需的最低核算价。
        """)

    st.stop()

# ── 文件类型检测 ──────────────────────────────────────────────────────────────
@st.cache_data
def detect_file_type(source):
    probe = pd.read_excel(source, engine="calamine", nrows=2)
    first_col = str(probe.columns[0]).lower()
    if "超过1w" in first_col or "超过" in first_col:
        return "store_raw"
    row1 = " ".join(str(v) for v in probe.iloc[0].values).lower()
    if "sku编码" in row1 and "折后单价" in row1:
        return "store_raw"
    all_cols = " ".join(str(c) for c in probe.columns)
    if "加权平均折后单价" in all_cols:
        return "agg"
    return "hq"

file_mode = detect_file_type(uploaded)

# ── 自动聚合（原始出库流水 → SKU级汇总）──────────────────────────────────────
@st.cache_data
def auto_aggregate(source):
    """将原始出库流水按 SKU/品牌 聚合，输出与 aggregate_outbound.py 一致的格式。"""
    probe = pd.read_excel(source, engine="calamine", nrows=3, header=None)
    hdr_row = 1
    for i in range(len(probe)):
        row_str = " ".join(str(v) for v in probe.iloc[i].values)
        if "品牌名称" in row_str and ("sku" in row_str.lower() or "编码" in row_str):
            hdr_row = i; break
    df = pd.read_excel(source, engine="calamine", skiprows=hdr_row, header=0)

    def fc(df, *kws):
        for kw in kws:
            if kw in df.columns: return kw           # 精确匹配优先
        for kw in kws:
            hits = [c for c in df.columns if kw in str(c)]
            if hits: return hits[0]
        return None

    sku_c    = fc(df, "康众sku编码", "sku编码", "SKU编码")
    name_c   = fc(df, "产品名称")
    brand_c  = fc(df, "品牌名称")
    cost_c   = fc(df, "成本价")       # 精确匹配→不会碰到"成本金额"
    settle_c = fc(df, "核算价")       # 精确匹配→不会碰到"促销核算价"
    sale_c   = fc(df, "折后单价")
    region_c = fc(df, "出库仓库所属大区", "大区")
    cat3_c   = fc(df, "产品分类描述", "三级类目名称")
    cat2_c   = fc(df, "二级类目名称")
    cat1_c   = fc(df, "一级类目名称")

    for col in [cost_c, settle_c, sale_c]:
        if col: df[col] = pd.to_numeric(df[col], errors="coerce")

    gkeys = [c for c in [sku_c, name_c, brand_c, cat3_c, cat2_c, cat1_c] if c]
    if not gkeys: return None

    spec = {"出库单量（单数）": (sku_c or gkeys[0], "count")}
    if cost_c:   spec["加权平均成本价"]   = (cost_c, "mean")
    if settle_c: spec["加权平均核算价"]   = (settle_c, "mean")
    if sale_c:   spec["加权平均折后单价"] = (sale_c, "mean")
    if region_c: spec["来源大区"]        = (region_c, lambda x: x.mode().iloc[0] if len(x) else "")

    result = df.groupby(gkeys, as_index=False).agg(**spec)

    rn = {}
    if sku_c    and sku_c    != "康众sku编码": rn[sku_c]    = "康众sku编码"
    if name_c   and name_c   != "产品名称":   rn[name_c]   = "产品名称"
    if brand_c  and brand_c  != "品牌名称":   rn[brand_c]  = "品牌名称"
    if rn: result = result.rename(columns=rn)
    if "加权平均核算价" in result and "加权平均折后单价" in result:
        result["折扣率（折后/核算）"] = result["加权平均折后单价"] / result["加权平均核算价"]
    return result

# ── HQ-only 文件（无门店折后价）→ 归一化为与 agg 模式兼容的 DataFrame ─────────
def find_col(cols, keywords, exclude=None):
    exclude = exclude or []
    for kw in keywords:
        for c in cols:
            c_lower = str(c).lower()
            if kw.lower() in c_lower and not any(ex.lower() in c_lower for ex in exclude):
                return c
    return None

def build_hq_agg(source):
    raw_df = pd.read_excel(source)
    cols = list(raw_df.columns)

    auto = {
        "SKU编码":  find_col(cols, ["sku编码","sku","康众编码","商品编码","库内编码","货号"]),
        "品牌":     find_col(cols, ["品牌"]),
        "品类":     find_col(cols, ["品类","产品分类","一级分类","商品目录","粗称","分类","类目名称","类目"]),
        "采购成本": find_col(cols, ["加权平均","加权成本","进货价","供应商价","采购成本"],
                             exclude=["核算价","金额"]),
        "核算价":   find_col(cols, ["核算价","采购核算","结算价","售价","零售价","销售价"],
                             exclude=["金额"]),
    }
    missing = [k for k, v in auto.items() if v is None]
    if missing:
        with st.expander("⚙️ 自动识别列失败，请手动选择（点击展开）", expanded=True):
            st.caption(f"未能自动识别：{', '.join(missing)}，请从下拉框选择对应列。")
            for field in missing:
                auto[field] = st.selectbox(f"{field} 对应哪一列？",
                                           ["（跳过）"] + cols, key=f"sel_{field}")
                if auto[field] == "（跳过）":
                    auto[field] = None

    if not auto["采购成本"] or not auto["核算价"]:
        st.warning("至少需要指定「采购成本」和「核算价」两列才能开始分析。")
        st.stop()

    keep = {v: k for k, v in auto.items() if v}
    hq = raw_df[list(keep.keys())].rename(columns=keep).copy()
    for col in ["SKU编码", "品牌", "品类"]:
        if col not in hq.columns:
            hq[col] = "—"
    hq["采购成本"] = pd.to_numeric(hq["采购成本"], errors="coerce")
    hq["核算价"]   = pd.to_numeric(hq["核算价"],   errors="coerce")

    valid = hq[hq["采购成本"].notna() & hq["核算价"].notna() &
               (hq["采购成本"] > 0) & (hq["核算价"] > 0)]
    if len(valid) > 0:
        cm, pm = valid["采购成本"].mean(), valid["核算价"].mean()
        if abs(cm - pm) / max(pm, 1) < 0.02:
            st.warning(
                f"⚠️ **检测到成本列与核算价列数值几乎相等（差异 < 2%）**\n\n"
                f"成本列 `{auto['采购成本']}` 均值 {cm:.2f}，核算价列 `{auto['核算价']}` 均值 {pm:.2f}。\n\n"
                "此文件可能是门店采购流水，采购核算价为门店从总部的拿货价，不含供应商进货成本，无法计算毛利率。"
            )
            st.stop()

    hq = hq.rename(columns={
        "SKU编码": "康众sku编码", "品牌": "品牌名称", "品类": "产品分类描述",
        "采购成本": "加权平均成本价", "核算价": "加权平均核算价",
    })
    return hq

# ── 三种模式统一归一化为 agg DataFrame，进入同一套四 Tab 分析 ─────────────────
@st.cache_data
def load_agg(source):
    df = pd.read_excel(source, engine="calamine")
    for c in ["加权平均成本价","加权平均核算价","加权平均折后单价","折扣率（折后/核算）"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df

if file_mode == "store_raw":
    with st.spinner("正在聚合出库流水……"):
        agg = auto_aggregate(uploaded)
    if agg is None:
        st.error("无法识别文件列结构，请确认文件包含「品牌名称」和「sku编码」列。")
        st.stop()
    n_raw_rows = len(pd.read_excel(uploaded, engine="calamine", skiprows=1, header=0))
    st.info(f"已自动聚合 {n_raw_rows:,} 条流水 → {len(agg):,} 个 SKU，直接进入分析。")
elif file_mode == "hq":
    agg = build_hq_agg(uploaded)
else:
    agg = load_agg(uploaded)

if True:  # 保留原有缩进层级，三种 file_mode 现共用以下全部逻辑
    agg = agg.copy()
    brand_col  = next((c for c in agg.columns if "品牌" in c), None)
    cat_col    = next((c for c in agg.columns if "分类" in c or "品类" in c), None)
    region_col = next((c for c in agg.columns if "大区" in c), None)
    name_col   = next((c for c in agg.columns if "产品名称" in c or c == "名称"), None)

    def sku_search_filter(df, search_text, sku_col):
        """按 SKU 编码 / 产品名称模糊匹配过滤（大小写不敏感）。"""
        if not search_text or not sku_col or sku_col not in df.columns:
            return df
        s = search_text.strip().lower()
        mask = df[sku_col].astype(str).str.lower().str.contains(s, na=False, regex=False)
        if name_col and name_col in df.columns:
            mask = mask | df[name_col].astype(str).str.lower().str.contains(s, na=False, regex=False)
        return df[mask]

    with st.sidebar:
        st.divider()
        st.header("筛选")
        _brands = ["全部"] + sorted(agg[brand_col].dropna().unique().tolist()) if brand_col else ["全部"]
        _cats   = sorted(agg[cat_col].dropna().unique().tolist()) if cat_col else []
        _sel_brand = st.selectbox("品牌", _brands)
        _sel_cat   = st.multiselect("品类", _cats, placeholder="全部（不限）") if _cats else []

    view = agg.copy()
    if _sel_brand != "全部" and brand_col:
        view = view[view[brand_col] == _sel_brand]
    if _sel_cat and cat_col:
        view = view[view[cat_col].isin(_sel_cat)]

    # 计算两层毛利率
    HAS_COST = "加权平均成本价" in view.columns and view["加权平均成本价"].notna().any()
    HAS_SALE = "加权平均折后单价" in view.columns and view["加权平均折后单价"].notna().any()

    # 双重噪音过滤：① 价格 ≥ 1元（赠品占位符） ② |毛利率| ≤ 500%（单位混淆/录入错误）
    MARGIN_CAP = 5.0

    hq_df = view[view["加权平均核算价"].notna() & view["加权平均成本价"].notna() &
                 (view["加权平均核算价"] >= 1) & (view["加权平均成本价"] > 0)].copy() if HAS_COST else pd.DataFrame()
    if len(hq_df):
        hq_df["总部毛利率"] = (hq_df["加权平均核算价"] - hq_df["加权平均成本价"]) / hq_df["加权平均核算价"]
        hq_df = hq_df[np.isfinite(hq_df["总部毛利率"]) & (hq_df["总部毛利率"].abs() <= MARGIN_CAP)].copy()

    st_df = view[view["加权平均核算价"].notna() & view["加权平均折后单价"].notna() &
                 (view["加权平均核算价"] >= 1) & (view["加权平均折后单价"] >= 1)].copy() if HAS_SALE else pd.DataFrame()
    if len(st_df):
        st_df["门店毛利率"] = (st_df["加权平均折后单价"] - st_df["加权平均核算价"]) / st_df["加权平均折后单价"]
        st_df = st_df[np.isfinite(st_df["门店毛利率"]) & (st_df["门店毛利率"].abs() <= MARGIN_CAP)].copy()

    tab1, tab2, tab3, tab4 = st.tabs(["📊 总部毛利分析", "🏪 门店定价分析", "🔗 综合对比", "💡 定价建议"])

    # ─ Tab1：总部毛利分析 ──────────────────────────────────────────────────────
    n_excl_hq = len(view[view["加权平均核算价"].notna() & (view["加权平均核算价"] < 1)]) if HAS_COST else 0
    with tab1:
        if not len(hq_df):
            st.warning("当前文件无法进行总部毛利分析（缺少加权平均成本价列）。")
        else:
            if n_excl_hq:
                st.caption(f"ℹ️ 已剔除 {n_excl_hq} 条核算价 < 1 元的异常记录（赠品/占位符），不参与分析。")
            thr1 = st.slider("毛利率基准线（%）", 5, 25, 10, 1, key="thr1")
            thr1f = thr1 / 100

            use_tiered1 = st.checkbox(
                "启用分层定价（按当前毛利率分层设定目标毛利率，只降不升）", value=False, key="tiered1")
            if use_tiered1:
                tc1, tc2, tc3 = st.columns(3)
                tier_high_floor1  = tc1.slider("高毛利层起点（%）", 15, 50, 20, 1, key="thf1")
                tier_high_target1 = tc2.slider(f"高毛利层（≥{tier_high_floor1}%）目标毛利率（%）",
                                                thr1, 40, 15, 1, key="tht1")
                tier_mid_target1  = tc3.slider(f"中毛利层（{thr1}%~{tier_high_floor1}%）目标毛利率（%）",
                                                thr1, 40, thr1 + 2, 1, key="tmt1")
                adj1 = 0
            else:
                adj1 = st.slider("调价幅度（%）", -30, 30, 0, 1, key="adj1",
                                 help="正数=涨价，负数=降价（仅对高于基准线的 SKU 生效）")

            d1 = hq_df.copy()
            if use_tiered1:
                high_floor1 = tier_high_floor1 / 100
                high_tgt1   = tier_high_target1 / 100
                mid_tgt1    = tier_mid_target1 / 100
                d1["调后核算价"] = np.select(
                    [d1["总部毛利率"] >= high_floor1,
                     (d1["总部毛利率"] > thr1f) & (d1["总部毛利率"] < high_floor1)],
                    [np.minimum(d1["加权平均成本价"] / (1 - high_tgt1), d1["加权平均核算价"]),
                     np.minimum(d1["加权平均成本价"] / (1 - mid_tgt1),  d1["加权平均核算价"])],
                    default=d1["加权平均核算价"],
                )
            else:
                d1["调后核算价"] = np.where(d1["总部毛利率"] > thr1f,
                                            d1["加权平均核算价"] * (1 + adj1/100), d1["加权平均核算价"])
            d1["调后毛利率"] = (d1["调后核算价"] - d1["加权平均成本价"]) / d1["调后核算价"]

            n_tot = len(d1); avg_b = d1["总部毛利率"].mean(); avg_a = d1["调后毛利率"].mean()
            n_red = int((d1["调后毛利率"] < 0.08).sum())
            n_yel = int(((d1["调后毛利率"] >= 0.08) & (d1["调后毛利率"] < thr1f)).sum())
            n_grn = int((d1["调后毛利率"] >= thr1f).sum())
            c1,c2,c3,c4 = st.columns(4)
            c1.metric("SKU 数", f"{n_tot:,}"); c2.metric("低于基准线", f"{n_red+n_yel:,}")
            c3.metric("当前均值毛利率", f"{avg_b:.1%}"); c4.metric("调整后均值", f"{avg_a:.1%}", f"{avg_a-avg_b:+.1%}")
            st.divider()
            v1,v2 = st.columns([1,2])
            with v1:
                st.markdown("#### SKU 健康分布")
                fig = go.Figure(go.Pie(labels=["健康","预警","危险"], values=[n_grn,n_yel,n_red],
                    hole=0.6, marker_colors=["#2ecc71","#f39c12","#e74c3c"], textinfo="percent+value",
                    hovertemplate="%{label}<br>%{value} 个<br>%{percent}<extra></extra>"))
                fig.update_layout(margin=dict(t=10,b=10,l=10,r=10), height=250,
                    legend=dict(orientation="h",yanchor="bottom",y=-0.3,font_size=11),
                    annotations=[dict(text=f"<b>{n_grn}</b><br>健康",x=0.5,y=0.5,font_size=14,showarrow=False)])
                st.plotly_chart(fig, use_container_width=True, config={"displayModeBar":False})
            with v2:
                if brand_col:
                    st.markdown("#### 各品牌均值毛利率")
                    b_agg = d1.groupby(brand_col)[["总部毛利率","调后毛利率"]].mean().sort_values("总部毛利率") * 100
                    fig2 = go.Figure()
                    fig2.add_trace(go.Bar(name="调价前", x=b_agg.index, y=b_agg["总部毛利率"],
                        marker_color="rgba(52,152,219,0.6)", hovertemplate="%{x}<br>%{y:.1f}%<extra>调价前</extra>"))
                    fig2.add_trace(go.Bar(name="调价后", x=b_agg.index, y=b_agg["调后毛利率"],
                        marker_color="rgba(46,204,113,0.6)", hovertemplate="%{x}<br>%{y:.1f}%<extra>调价后</extra>"))
                    fig2.add_hline(y=thr1, line_dash="dash", line_color="#e74c3c", line_width=1.5)
                    fig2.update_layout(margin=dict(t=10,b=10,l=10,r=10), height=250, barmode="group",
                        legend=dict(orientation="h",yanchor="bottom",y=-0.35,font_size=11))
                    st.plotly_chart(fig2, use_container_width=True, config={"displayModeBar":False})

            # 多幅度调价对比（分层模式下每个 SKU 已有精确落地价，多幅度对比无意义，隐藏）
            if not use_tiered1:
                st.divider()
                st.markdown("#### 多幅度调价对比")
                scen_rows1 = []
                for s in [-10, -5, 0, 5, 10]:
                    adj_price1 = np.where(d1["总部毛利率"] > thr1f,
                                           d1["加权平均核算价"] * (1 + s/100), d1["加权平均核算价"])
                    adj_margin1 = (adj_price1 - d1["加权平均成本价"]) / adj_price1
                    scen_rows1.append({
                        "调价幅度": f"{s:+d}%",
                        "均值毛利率": adj_margin1.mean(),
                        f"低于{thr1}%的SKU数": int((adj_margin1 < thr1f).sum()),
                        "危险SKU数(<8%)": int((adj_margin1 < 0.08).sum()),
                        "健康SKU数": int((adj_margin1 >= thr1f).sum()),
                    })
                scen_df1 = pd.DataFrame(scen_rows1)
                def _hl_scen1(row):
                    if row["调价幅度"] == f"{adj1:+d}%":
                        return ["background-color:#1a4a2e;color:#fff;font-weight:bold"] * len(row)
                    return [""] * len(row)
                st.dataframe(scen_df1.style.apply(_hl_scen1, axis=1).format({"均值毛利率":"{:.1%}"}),
                             use_container_width=True, hide_index=True)

            # 品牌×品类 毛利率热力图（至少各 2 个才有意义）
            if brand_col and cat_col and d1[brand_col].nunique() >= 2 and d1[cat_col].nunique() >= 2:
                st.divider()
                st.markdown("#### 品牌 × 品类 毛利率热力图（调价后）")
                pivot1 = d1.groupby([cat_col, brand_col])["调后毛利率"].mean().unstack(fill_value=np.nan)
                z1 = pivot1.values * 100
                txt1 = [[f"{v:.1f}%" if not np.isnan(v) else "—" for v in row] for row in z1]
                hm1 = go.Figure(go.Heatmap(
                    z=z1, x=list(pivot1.columns), y=list(pivot1.index),
                    colorscale="RdYlGn", zmin=0, zmax=30, text=txt1, texttemplate="%{text}",
                    hovertemplate="品类：%{y}<br>品牌：%{x}<br>均值毛利率：%{z:.1f}%<extra></extra>",
                    colorbar=dict(title="毛利率%")))
                hm1.update_layout(margin=dict(t=10,b=10,l=10,r=10),
                    height=max(260, len(pivot1.index)*32 + 80),
                    xaxis=dict(tickangle=-20, side="top"),
                    plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)")
                st.plotly_chart(hm1, use_container_width=True, config={"displayModeBar":False})

            st.divider()
            st.markdown("#### SKU 明细")
            sku_col1 = next((c for c in d1.columns if "sku" in c.lower() or "编码" in c), None)
            search1 = st.text_input("🔍 搜索 SKU 编码 / 品名", key="search1", placeholder="输入关键字过滤下表")
            _cols1 = [c for c in [sku_col1, brand_col, cat_col, "加权平均成本价",
                                   "加权平均核算价","总部毛利率","调后核算价","调后毛利率"] if c and c in d1.columns]
            show1 = d1[_cols1].copy() if _cols1 else d1
            show1 = sku_search_filter(show1, search1, sku_col1)
            def _sm1(v):
                try:
                    f=float(v); return ("background-color:#f8d7da" if f<0.08 else
                        "background-color:#fff3cd" if f<thr1f else "background-color:#d4edda") + ";color:#1a1a1a"
                except: return ""
            subset1 = [c for c in ["总部毛利率","调后毛利率"] if c in show1.columns]
            fmt1 = {c:"{:.1%}" for c in subset1}
            fmt1.update({c:"{:.2f}" for c in ["加权平均成本价","加权平均核算价","调后核算价"] if c in show1.columns})
            st.dataframe(show1.style.map(_sm1, subset=subset1).format(fmt1),
                         use_container_width=True, height=380, hide_index=True)

    # ─ Tab2：门店定价分析 ──────────────────────────────────────────────────────
    with tab2:
        if not len(st_df):
            st.warning("当前文件缺少加权平均折后单价列，无法进行门店定价分析。")
        else:
            thr2 = st.slider("门店毛利率基准线（%）", 0, 30, 10, 1, key="thr2")
            thr2f = thr2 / 100
            n2_loss = int((st_df["门店毛利率"] < 0).sum())
            n2_low  = int(((st_df["门店毛利率"] >= 0) & (st_df["门店毛利率"] < thr2f)).sum())
            n2_ok   = int((st_df["门店毛利率"] >= thr2f).sum())
            avg2 = st_df["门店毛利率"].mean()
            st.info(f"共 **{len(st_df):,}** 条 SKU · 亏本销售 **{n2_loss}** 条 · 平均门店毛利率 **{avg2:.1%}**")
            c1,c2,c3,c4 = st.columns(4)
            c1.metric("SKU 数", f"{len(st_df):,}")
            c2.metric("亏本", f"{n2_loss}", delta_color="inverse")
            c3.metric("低于基准线", f"{n2_low}", delta_color="inverse")
            c4.metric("平均门店毛利率", f"{avg2:.1%}")
            st.divider()
            v1,v2 = st.columns([1,2])
            with v1:
                st.markdown("#### 门店定价健康分布")
                fig = go.Figure(go.Pie(labels=["健康","微利","亏本"], values=[n2_ok,n2_low,n2_loss],
                    hole=0.6, marker_colors=["#2ecc71","#f39c12","#e74c3c"], textinfo="percent+value"))
                fig.update_layout(margin=dict(t=10,b=10,l=10,r=10), height=250,
                    legend=dict(orientation="h",yanchor="bottom",y=-0.3,font_size=11),
                    annotations=[dict(text=f"<b>{n2_ok}</b><br>健康",x=0.5,y=0.5,font_size=14,showarrow=False)])
                st.plotly_chart(fig, use_container_width=True, config={"displayModeBar":False})
            with v2:
                if brand_col:
                    st.markdown("#### 各品牌平均门店毛利率")
                    b2 = st_df.groupby(brand_col)["门店毛利率"].mean().sort_values() * 100
                    bar2 = go.Figure(go.Bar(x=b2.values, y=b2.index, orientation="h",
                        marker_color=["#e74c3c" if v<thr2 else "#2ecc71" for v in b2.values],
                        hovertemplate="%{y}<br>%{x:.1f}%<extra></extra>"))
                    bar2.add_vline(x=thr2, line_dash="dash", line_color="#e74c3c")
                    bar2.update_layout(margin=dict(t=10,b=10,l=10,r=10), height=250, xaxis_title="门店毛利率（%）")
                    st.plotly_chart(bar2, use_container_width=True, config={"displayModeBar":False})
            st.divider()
            st.markdown("#### 低于基准线 SKU 明细")
            sku_col2 = next((c for c in st_df.columns if "sku" in c.lower() or "编码" in c), None)
            search2 = st.text_input("🔍 搜索 SKU 编码 / 品名", key="search2", placeholder="输入关键字过滤下表")
            prob2 = st_df[st_df["门店毛利率"] < thr2f].sort_values("门店毛利率")
            prob2 = sku_search_filter(prob2, search2, sku_col2)
            cols2 = [c for c in [sku_col2, brand_col, cat_col, "加权平均核算价","加权平均折后单价","门店毛利率","折扣率（折后/核算）"] if c and c in prob2.columns]
            def _sm2(v):
                try:
                    f=float(v); return ("background-color:#f8d7da" if f<0 else "background-color:#fff3cd") + ";color:#1a1a1a"
                except: return ""
            fmt2 = {c:"{:.1%}" for c in ["门店毛利率","折扣率（折后/核算）"] if c in cols2}
            fmt2.update({c:"{:.2f}" for c in ["加权平均核算价","加权平均折后单价"] if c in cols2})
            st.dataframe(prob2[cols2].style.map(_sm2, subset=["门店毛利率"]).format(fmt2),
                         use_container_width=True, height=380, hide_index=True)
            st.download_button("📋 下载低于基准线明细（CSV）",
                data=prob2[cols2].to_csv(index=False).encode("utf-8-sig"),
                file_name=f"门店定价异常_{thr2}pct基准线.csv", mime="text/csv")

    # ─ Tab3：综合对比 ──────────────────────────────────────────────────────────
    with tab3:
        if not len(hq_df) or not len(st_df):
            st.warning("综合对比需要同时具备成本价和折后单价数据。")
        else:
            st.markdown("#### 价值链全链路毛利率对比")
            st.caption("总部毛利率 = (核算价 - 成本价) / 核算价 ｜ 门店毛利率 = (折后单价 - 核算价) / 折后单价 ｜ 全链路毛利率 = (折后单价 - 成本价) / 折后单价")

            # 合并两层数据（按 SKU 编码 join）
            sku_col = next((c for c in agg.columns if "sku" in c.lower() or "编码" in c), None)
            if sku_col:
                merged = pd.merge(
                    hq_df[[sku_col, brand_col or "品牌名称", "加权平均成本价","加权平均核算价","总部毛利率"]],
                    st_df[[sku_col, "加权平均折后单价","门店毛利率"]],
                    on=sku_col, how="inner"
                )
            else:
                merged = pd.concat([hq_df.assign(门店毛利率=None), st_df.assign(总部毛利率=None)], ignore_index=True)

            if len(merged):
                merged["全链路毛利率"] = (merged["加权平均折后单价"] - merged["加权平均成本价"]) / merged["加权平均折后单价"]
                merged["⚠️"] = np.where(
                    (merged["总部毛利率"] > 0.1) & (merged["门店毛利率"] < 0),
                    "总部高利润但门店亏本", np.where(
                    (merged["门店毛利率"] > 0.1) & (merged["总部毛利率"] < 0.08),
                    "门店利润健康但总部偏低", ""))

                # 品牌汇总图
                if brand_col and brand_col in merged.columns:
                    b3 = merged.groupby(brand_col)[["总部毛利率","门店毛利率","全链路毛利率"]].mean().sort_values("全链路毛利率") * 100
                    fig3 = go.Figure()
                    for col_name, color, label in [
                        ("总部毛利率","rgba(52,152,219,0.8)","总部毛利率"),
                        ("门店毛利率","rgba(231,76,60,0.8)","门店毛利率"),
                        ("全链路毛利率","rgba(46,204,113,0.8)","全链路毛利率"),
                    ]:
                        if col_name in b3.columns:
                            fig3.add_trace(go.Bar(name=label, x=b3.index, y=b3[col_name],
                                marker_color=color, hovertemplate=f"%{{x}}<br>{label} %{{y:.1f}}%<extra></extra>"))
                    fig3.add_hline(y=0, line_color="#888", line_width=1)
                    fig3.update_layout(barmode="group", height=320,
                        margin=dict(t=10,b=10,l=10,r=10),
                        legend=dict(orientation="h",yanchor="bottom",y=-0.35,font_size=11))
                    st.plotly_chart(fig3, use_container_width=True, config={"displayModeBar":False})

                st.divider()
                c1,c2,c3 = st.columns(3)
                c1.metric("平均总部毛利率", f"{merged['总部毛利率'].mean():.1%}")
                c2.metric("平均门店毛利率", f"{merged['门店毛利率'].mean():.1%}")
                c3.metric("平均全链路毛利率", f"{merged['全链路毛利率'].mean():.1%}")

                st.divider()
                flag_cnt = int((merged["⚠️"] != "").sum())
                if flag_cnt:
                    st.warning(f"⚠️ 发现 **{flag_cnt}** 条异常：总部高毛利但门店亏本，或反向不匹配。")
                show3_cols = [c for c in [sku_col, brand_col, "加权平均成本价","加权平均核算价",
                    "加权平均折后单价","总部毛利率","门店毛利率","全链路毛利率","⚠️"] if c and c in merged.columns]
                fmt3 = {c:"{:.1%}" for c in ["总部毛利率","门店毛利率","全链路毛利率"] if c in show3_cols}
                fmt3.update({c:"{:.2f}" for c in ["加权平均成本价","加权平均核算价","加权平均折后单价"] if c in show3_cols})

                def _flag(row):
                    if row.get("⚠️",""):
                        return ["background-color:#fff0e0;color:#1a1a1a"] * len(row)
                    return [""] * len(row)

                st.markdown("#### SKU 全链路明细")
                search3 = st.text_input("🔍 搜索 SKU 编码 / 品名", key="search3", placeholder="输入关键字过滤下表")
                show3 = sku_search_filter(merged[show3_cols], search3, sku_col)
                st.dataframe(show3.sort_values("全链路毛利率")
                    .style.apply(_flag, axis=1).format(fmt3),
                    use_container_width=True, height=420, hide_index=True)
                st.download_button("📥 下载综合对比明细（CSV）",
                    data=show3.to_csv(index=False).encode("utf-8-sig"),
                    file_name="综合对比全链路.csv", mime="text/csv")
            else:
                st.info("没有同时具备三层价格的 SKU，无法生成综合对比。")

    # ─ Tab4：定价建议 ──────────────────────────────────────────────────────────
    with tab4:
        if not len(hq_df) or not len(st_df):
            st.warning("定价建议需要同时具备总部毛利率和门店毛利率数据。")
        else:
            st.caption("以 SKU 为维度，综合总部利润 × 门店利润 × 销量三者，输出核算价调整方向、建议新价格与业务影响测算")

            col_s1, col_s2 = st.columns(2)
            hq_thr4 = col_s1.slider("总部毛利率基准线（%）", 0, 20, 8, 1, key="hq_thr4")
            st_thr4 = col_s2.slider("门店毛利率基准线（%）", 0, 30, 15, 1, key="st_thr4")
            hq_f4 = hq_thr4 / 100
            st_f4 = st_thr4 / 100

            sku_c4 = next((c for c in hq_df.columns if "sku" in c.lower() or "编码" in c), None)
            vol_c  = next((c for c in view.columns if "出库单量" in c or "单数" in c), None)
            if not sku_c4:
                st.warning("无法识别 SKU 编码列。")
            else:
                left_c  = [c for c in [sku_c4, brand_col, cat_col, "加权平均成本价", "加权平均核算价", "总部毛利率"] if c and c in hq_df.columns]
                right_c = [c for c in [sku_c4, "加权平均折后单价", "门店毛利率"] if c and c in st_df.columns]
                rec = pd.merge(hq_df[left_c], st_df[right_c], on=sku_c4, how="inner")

                # 报告因单侧缺数据被排除的 SKU 数量，避免定价建议清单静默丢数据
                n_only_hq = hq_df[sku_c4].nunique() - rec[sku_c4].nunique()
                n_only_st = st_df[sku_c4].nunique() - rec[sku_c4].nunique()
                if n_only_hq or n_only_st:
                    st.caption(
                        f"ℹ️ 本次定价建议基于 **{rec[sku_c4].nunique():,}** 个同时具备总部+门店数据的 SKU；"
                        f"另有 **{n_only_hq:,}** 个仅有总部数据（缺门店折后价）、"
                        f"**{n_only_st:,}** 个仅有门店数据（缺总部成本价），未纳入本建议，"
                        "可分别在 Tab1 / Tab2 中查看。"
                    )

                if vol_c and vol_c in view.columns:
                    rec = pd.merge(rec, view[[sku_c4, vol_c]].drop_duplicates(), on=sku_c4, how="left")
                    rec[vol_c] = rec[vol_c].fillna(1)
                else:
                    vol_c = "出库单量（单数）"
                    rec[vol_c] = 1

                # 分类 & 建议新价格
                def _classify(row):
                    hq_ok = row["总部毛利率"] >= hq_f4
                    st_ok = row["门店毛利率"] >= st_f4
                    K = row["加权平均核算价"]
                    C = row["加权平均成本价"]
                    P = row["加权平均折后单价"]
                    vol = row[vol_c]

                    if hq_ok and not st_ok:
                        K_new = max(P * (1 - st_f4), C * 1.01)
                        delta  = (K_new - K) / K
                        profit_chg = (K_new - C - (K - C)) * vol
                        return ("A", "建议降核算价", round(K_new, 2), f"{delta:+.1%}", round(profit_chg, 0),
                                f"总部毛利率 {row['总部毛利率']:.1%} 充足，但门店毛利率仅 {row['门店毛利率']:.1%}（<{st_thr4}% 基准）；"
                                f"建议核算价从 {K:.2f} 降至 {K_new:.2f}，门店毛利率恢复至 {st_thr4}%")

                    elif not hq_ok and st_ok:
                        K_new = C / (1 - hq_f4) if (1 - hq_f4) > 0 else K
                        K_new = min(K_new, P * (1 - st_f4 * 0.5))
                        delta  = (K_new - K) / K
                        profit_chg = (K_new - C - (K - C)) * vol
                        return ("B", "可升核算价", round(K_new, 2), f"{delta:+.1%}", round(profit_chg, 0),
                                f"门店毛利率 {row['门店毛利率']:.1%} 较高，总部毛利率仅 {row['总部毛利率']:.1%}（<{hq_thr4}% 基准）；"
                                f"建议核算价从 {K:.2f} 提至 {K_new:.2f}，总部毛利率恢复至 {hq_thr4}%")

                    elif not hq_ok and not st_ok:
                        K_new = C / (1 - hq_f4) if (1 - hq_f4) > 0 else K
                        delta  = (K_new - K) / K
                        profit_chg = (K_new - C - (K - C)) * vol
                        return ("C", "全链路亏损，审视成本", round(K_new, 2),
                                f"{delta:+.1%}（仅覆盖总部基准）", round(profit_chg, 0),
                                f"总部 {row['总部毛利率']:.1%}、门店 {row['门店毛利率']:.1%} 均低于基准；"
                                f"建议核算价最低调至 {K_new:.2f} 以覆盖总部成本，需同步评估供应商价格与市场竞争力")

                    else:
                        return ("D", "定价合理，维持", round(K, 2), "—", 0.0,
                                f"总部 {row['总部毛利率']:.1%}、门店 {row['门店毛利率']:.1%} 均达标，无需调整")

                rec[["类型","建议","建议新核算价","调整幅度","总部毛利额变化（元）","原因"]] =                     rec.apply(_classify, axis=1, result_type="expand")

                # 优先级：类型 × 销量分层
                q67 = rec[vol_c].quantile(0.67)
                q33 = rec[vol_c].quantile(0.33)
                def _tier(v): return "高" if v >= q67 else ("中" if v >= q33 else "低")
                rec["销量层级"] = rec[vol_c].apply(_tier)

                def _pri(row):
                    t = row["类型"]; tier = row["销量层级"]
                    if t in ("A","B") and tier == "高": return "🔥 高"
                    if (t in ("A","B") and tier in ("中","低")) or (t == "C" and tier == "高"): return "⚠️ 中"
                    return "📌 低"
                rec["优先级"] = rec.apply(_pri, axis=1)

                pmap = {"🔥 高": 0, "⚠️ 中": 1, "📌 低": 2}
                rec["_p"] = rec["优先级"].map(pmap)
                rec = rec.sort_values(["_p", vol_c], ascending=[True, False]).drop(columns="_p")

                # 汇总
                na = int((rec["类型"]=="A").sum()); nb = int((rec["类型"]=="B").sum())
                nc = int((rec["类型"]=="C").sum()); nd = int((rec["类型"]=="D").sum())
                c1,c2,c3,c4_ = st.columns(4)
                c1.metric("🔴 A — 建议降核算价", na, f"影响 {int(rec[rec['类型']=='A'][vol_c].sum()):,} 单")
                c2.metric("🟢 B — 可升核算价",   nb, f"影响 {int(rec[rec['类型']=='B'][vol_c].sum()):,} 单")
                c3.metric("🟡 C — 全链路亏损",   nc, f"影响 {int(rec[rec['类型']=='C'][vol_c].sum()):,} 单")
                c4_.metric("⚪ D — 定价合理",    nd)

                total_chg = rec[rec["类型"].isin(["A","B","C"])]["总部毛利额变化（元）"].sum()
                st.info(f"按建议全部调价后，总部毛利额预计变化：**{total_chg:+,.0f} 元**（基于当前出库单量测算）")
                st.divider()

                sel_types = st.multiselect(
                    "筛选建议类型",
                    ["A 降核算价","B 可升核算价","C 全链路亏损","D 定价合理"],
                    default=["A 降核算价","B 可升核算价","C 全链路亏损"])
                tm = {"A 降核算价":"A","B 可升核算价":"B","C 全链路亏损":"C","D 定价合理":"D"}
                disp4 = rec[rec["类型"].isin([tm[t] for t in sel_types])].copy() if sel_types else rec.copy()

                search4 = st.text_input("🔍 搜索 SKU 编码 / 品名", key="search4", placeholder="输入关键字过滤下表")
                disp4 = sku_search_filter(disp4, search4, sku_c4)

                show_c4 = [c for c in [
                    sku_c4, brand_col, cat_col,
                    "加权平均成本价","加权平均核算价","建议新核算价","调整幅度",
                    "加权平均折后单价","总部毛利率","门店毛利率",
                    vol_c,"销量层级","优先级","类型","建议","总部毛利额变化（元）","原因"
                ] if c and c in disp4.columns]

                tc = {"A":"#f8d7da","B":"#d4edda","C":"#fff3cd","D":"#e9ecef"}
                def _cr4(row): return [f"background-color:{tc.get(row['类型'],'')};color:#1a1a1a"]*len(row)
                fmt4 = {c:"{:.1%}" for c in ["总部毛利率","门店毛利率"] if c in show_c4}
                fmt4.update({c:"{:.2f}" for c in ["加权平均成本价","加权平均核算价","建议新核算价","加权平均折后单价"] if c in show_c4})
                fmt4["总部毛利额变化（元）"] = "{:+,.0f}"

                st.dataframe(disp4[show_c4].style.apply(_cr4, axis=1).format(fmt4, na_rep="—"),
                             use_container_width=True, height=480, hide_index=True)

                def build_excel(df, hq_thr, st_thr):
                    from io import BytesIO
                    import math
                    from openpyxl import Workbook
                    from openpyxl.styles import PatternFill, Font, Alignment
                    from openpyxl.utils import get_column_letter

                    wb = Workbook()
                    ws = wb.active; ws.title = "定价建议"
                    ws.freeze_panes = "A2"

                    fills = {"A": PatternFill("solid", fgColor="F8D7DA"),
                             "B": PatternFill("solid", fgColor="D4EDDA"),
                             "C": PatternFill("solid", fgColor="FFF3CD"),
                             "D": PatternFill("solid", fgColor="E9ECEF")}
                    hdr_fill  = PatternFill("solid", fgColor="2C3E50")
                    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
                    pct_set   = {"总部毛利率","门店毛利率"}
                    price_set = {"加权平均成本价","加权平均核算价","建议新核算价","加权平均折后单价"}
                    profit_set= {"总部毛利额变化（元）"}

                    cols = list(df.columns)
                    for j, h in enumerate(cols, 1):
                        cell = ws.cell(row=1, column=j, value=h)
                        cell.fill = hdr_fill; cell.font = hdr_font
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws.row_dimensions[1].height = 30

                    type_idx = cols.index("类型") if "类型" in cols else -1
                    for i, row_vals in enumerate(df.values, 2):
                        rtype = str(row_vals[type_idx]) if type_idx >= 0 else "D"
                        fill  = fills.get(rtype, PatternFill())
                        for j, (col, val) in enumerate(zip(cols, row_vals), 1):
                            v = None if (isinstance(val, float) and math.isnan(val)) else val
                            cell = ws.cell(row=i, column=j, value=v)
                            cell.fill = fill
                            cell.alignment = Alignment(vertical="center", wrap_text=(col in ("原因","调整幅度")))
                            if col in pct_set and isinstance(v, (int, float)):
                                cell.number_format = "0.0%"
                            elif col in price_set and isinstance(v, (int, float)):
                                cell.number_format = "#,##0.00"
                            elif col in profit_set and isinstance(v, (int, float)):
                                cell.number_format = '+#,##0;-#,##0;"-"'

                    for j, col in enumerate(cols, 1):
                        w = max(len(str(col)), max((len(str(v)) for v in df[col].values if v is not None), default=0))
                        ws.column_dimensions[get_column_letter(j)].width = min(w * 1.2 + 2, 52)

                    ws2 = wb.create_sheet("分类说明")
                    legend = [
                        ("类型","建议方向","判断条件","优先级逻辑","颜色"),
                        ("A","建议降核算价", f"总部毛利≥{hq_thr}% 且 门店毛利<{st_thr}%",
                         "A+高销量→🔥高；A+中低→⚠️中","红"),
                        ("B","可升核算价",   f"总部毛利<{hq_thr}% 且 门店毛利≥{st_thr}%",
                         "B+高销量→🔥高；B+中低→⚠️中","绿"),
                        ("C","全链路亏损",   f"总部毛利<{hq_thr}% 且 门店毛利<{st_thr}%",
                         "C+高销量→⚠️中；其余→📌低","黄"),
                        ("D","定价合理维持", f"总部毛利≥{hq_thr}% 且 门店毛利≥{st_thr}%",
                         "全部→📌低","灰"),
                        ("","建议新核算价公式",
                         "A类：折后单价×(1-门店基准)；B/C类：成本价÷(1-总部基准)","",""),
                        ("","业务影响公式",
                         "总部毛利额变化 = (新核算价-成本价 - 旧毛利额) × 出库单量","",""),
                    ]
                    for i, row in enumerate(legend, 1):
                        for j, v in enumerate(row, 1):
                            cell = ws2.cell(row=i, column=j, value=v)
                            if i == 1: cell.font = Font(bold=True)
                            elif j == 1 and row[0] in fills: cell.fill = fills[row[0]]
                    for j in range(1, 6):
                        ws2.column_dimensions[get_column_letter(j)].width = 40

                    buf = BytesIO(); wb.save(buf); buf.seek(0)
                    return buf.getvalue()

                st.download_button(
                    "📥 下载定价建议（带格式 Excel）",
                    data=build_excel(disp4[show_c4], hq_thr4, st_thr4),
                    file_name=f"核算价调整建议_{hq_thr4}pct总部_{st_thr4}pct门店.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
